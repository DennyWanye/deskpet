# SPDX-FileCopyrightText: 2026 DennyWanye
# SPDX-License-Identifier: BUSL-1.1

"""Excel generation tool — openpyxl wrapper (beta builtin skill).

The LLM produces a structured ``spec`` (a dict, or a JSON string of one);
this module turns it into a real ``.xlsx`` on disk. Mirrors the design of
:mod:`ppt_tools`: the model decides *content + structure*, this module
owns the *rendering*.

Spec shape::

    {
      "sheets": [
        {
          "name": "开支",
          "rows": [["类别","金额"], ["餐饮", 1200], ["合计", "=SUM(B2:B2)"]],
          "header_row": true,
          "column_widths": {"A": 18, "B": 12},
          "freeze_panes": "A2",
          "conditional_format": {"range": "B2:B3", "rule": "color_scale"},
          "chart": {"type": "bar", "title": "开支",
                    "data": "B1:B3", "categories": "A2:A3", "anchor": "D2"}
        }
      ]
    }

* Any cell whose value is a string starting with ``=`` is written as a
  live formula (openpyxl handles this natively).
* ``header_row`` styles row 1 bold with a fill + autofits columns.
* ``conditional_format.rule`` ∈ {color_scale, data_bar, greater_than}.
* ``chart.type`` ∈ {bar, line, pie}.

Failure philosophy: a bad spec returns ``{"ok": false, "error": ...}`` —
never raises out of the handler. openpyxl missing → same.
"""
from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any, Optional

from . import office_paths

log = logging.getLogger(__name__)

try:  # openpyxl is a hard dep for this skill, but degrade gracefully.
    from openpyxl import Workbook
    from openpyxl.cell.cell import Cell
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except Exception:  # noqa: BLE001
    _HAS_OPENPYXL = False

try:  # image embedding needs Pillow under the hood; degrade if absent.
    from openpyxl.drawing.image import Image as _XLImage

    _HAS_XLIMAGE = True
except Exception:  # noqa: BLE001
    _HAS_XLIMAGE = False


_HEADER_FILL = "FF1F4E78"
_HEADER_FONT = "FFFFFFFF"


def _coerce_spec(spec: Any) -> Optional[dict[str, Any]]:
    """Accept a dict or a JSON string; return a dict or None."""
    if isinstance(spec, dict):
        return spec
    if isinstance(spec, str):
        try:
            parsed = json.loads(spec)
        except json.JSONDecodeError:
            return None
        return parsed if isinstance(parsed, dict) else None
    return None


def _apply_header_style(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    font = Font(bold=True, color=_HEADER_FONT)
    align = Alignment(horizontal="center", vertical="center")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _autofit_columns(ws, rows: list[list[Any]]) -> None:
    """Best-effort column auto-width from the longest cell in each column."""
    if not rows:
        return
    ncols = max((len(r) for r in rows), default=0)
    for c in range(ncols):
        longest = 0
        for r in rows:
            if c < len(r) and r[c] is not None:
                # CJK chars render ~2x wide; weight them.
                text = str(r[c])
                width = sum(2 if ord(ch) > 0x2E7F else 1 for ch in text)
                longest = max(longest, width)
        ws.column_dimensions[get_column_letter(c + 1)].width = min(
            max(longest + 2, 8), 60
        )


def _apply_conditional_format(ws, cf: dict[str, Any]) -> None:
    rng = cf.get("range")
    rule = (cf.get("rule") or "color_scale").lower()
    if not rng:
        return
    if rule == "color_scale":
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="FFF8696B",
                mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
                end_type="max", end_color="FF63BE7B",
            ),
        )
    elif rule == "data_bar":
        ws.conditional_formatting.add(
            rng,
            DataBarRule(start_type="min", end_type="max", color="FF638EC6"),
        )
    elif rule == "greater_than":
        threshold = cf.get("value", 0)
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThan",
                formula=[str(threshold)],
                fill=PatternFill("solid", fgColor="FFFFC7CE"),
            ),
        )


def _add_chart(ws, chart_spec: dict[str, Any]) -> None:
    ctype = (chart_spec.get("type") or "bar").lower()
    if ctype == "line":
        chart = LineChart()
    elif ctype == "pie":
        chart = PieChart()
    else:
        chart = BarChart()
    chart.title = chart_spec.get("title") or None
    data_rng = chart_spec.get("data")
    cats_rng = chart_spec.get("categories")
    if data_rng:
        chart.add_data(
            Reference(ws, range_string=f"{ws.title}!{data_rng}"),
            titles_from_data=True,
        )
    if cats_rng:
        chart.set_categories(Reference(ws, range_string=f"{ws.title}!{cats_rng}"))
    anchor = chart_spec.get("anchor") or "H2"
    ws.add_chart(chart, anchor)


def _hex_argb(c: Any) -> Optional[str]:
    """'#1F4E78' / '1F4E78' / 'FF1F4E78' → 8-hex ARGB; bad input → None."""
    if not c:
        return None
    s = str(c).lstrip("#").upper()
    if len(s) == 6:
        return "FF" + s
    if len(s) == 8:
        return s
    return None


def _iter_cells(ws, rng: str):
    """Flatten ws[rng] into a flat cell iterator (handles single cell,
    a 1-D row/col tuple, and a 2-D tuple-of-tuples uniformly)."""
    try:
        sel = ws[rng]
    except Exception:  # noqa: BLE001 — bad range string
        return
    if isinstance(sel, Cell):
        yield sel
        return
    for item in sel:
        if isinstance(item, Cell):
            yield item
        else:  # nested tuple (multi-row range)
            for cell in item:
                yield cell


def _thin_border() -> "Border":
    side = Side(style="thin", color="FFBFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def _apply_number_formats(ws, specs: Any, nrows: int) -> None:
    """specs: list of {range:'B2:B9', format:'#,##0.00'} or
    {col:'B', format:'0.0%', from_row?:2}. Column form applies to data
    rows (default from row 2, i.e. skipping a header)."""
    if not isinstance(specs, list):
        return
    for spec in specs:
        if not isinstance(spec, dict):
            continue
        fmt = spec.get("format")
        if not fmt:
            continue
        rng = spec.get("range")
        col = spec.get("col")
        if rng:
            for cell in _iter_cells(ws, str(rng)):
                cell.number_format = str(fmt)
        elif col:
            start = int(spec.get("from_row", 2) or 2)
            for r in range(start, nrows + 1):
                ws[f"{col}{r}"].number_format = str(fmt)


def _apply_cell_styles(ws, styles: Any) -> None:
    """styles: list of {range, bold?, italic?, font_color?, fill?, align?,
    font_size?, border?, wrap?}. Each applies to every cell in the range."""
    if not isinstance(styles, list):
        return
    for st in styles:
        if not isinstance(st, dict) or not st.get("range"):
            continue
        font_kw: dict[str, Any] = {}
        if st.get("bold"):
            font_kw["bold"] = True
        if st.get("italic"):
            font_kw["italic"] = True
        fc = _hex_argb(st.get("font_color"))
        if fc:
            font_kw["color"] = fc
        if st.get("font_size"):
            try:
                font_kw["size"] = float(st["font_size"])
            except (ValueError, TypeError):
                pass
        fill = _hex_argb(st.get("fill"))
        align = st.get("align")
        wrap = bool(st.get("wrap"))
        border = _thin_border() if st.get("border") else None
        for cell in _iter_cells(ws, str(st["range"])):
            if font_kw:
                cell.font = Font(**font_kw)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            if align in ("left", "center", "right") or wrap:
                cell.alignment = Alignment(
                    horizontal=align if align in ("left", "center", "right") else None,
                    vertical="center",
                    wrap_text=wrap,
                )
            if border is not None:
                cell.border = border


def _apply_merges(ws, merges: Any) -> None:
    """merges: list of A1 ranges ('A1:C1'). Excel keeps the top-left
    cell's value; we merge after rows/styles so the anchor is styled."""
    if not isinstance(merges, list):
        return
    for m in merges:
        try:
            ws.merge_cells(str(m))
        except Exception as exc:  # noqa: BLE001 — bad range is non-fatal
            log.warning("excel merge skipped %r: %s", m, exc)


def _add_images(ws, images: Any) -> None:
    """images: list of {path, anchor?, width?, height?}. Needs Pillow."""
    if not isinstance(images, list):
        return
    if not _HAS_XLIMAGE:
        log.warning("excel images skipped: Pillow/openpyxl image support missing")
        return
    for im in images:
        if not isinstance(im, dict):
            continue
        p = im.get("path")
        if not p or not Path(str(p)).exists():
            log.warning("excel image path missing: %r", p)
            continue
        try:
            img = _XLImage(str(p))
            if im.get("width"):
                img.width = int(im["width"])
            if im.get("height"):
                img.height = int(im["height"])
            ws.add_image(img, str(im.get("anchor") or "H2"))
        except Exception as exc:  # noqa: BLE001 — image is non-critical
            log.warning("excel image skipped: %s", exc)


def _build_sheet(wb, sheet_spec: dict[str, Any], first: bool) -> None:
    name = str(sheet_spec.get("name") or "Sheet")[:31]  # Excel 31-char cap
    if first:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(title=name)

    rows = sheet_spec.get("rows") or []
    if not isinstance(rows, list):
        rows = []
    for r in rows:
        ws.append(list(r) if isinstance(r, (list, tuple)) else [r])

    ncols = max((len(r) for r in rows if isinstance(r, (list, tuple))), default=0)

    if sheet_spec.get("header_row") and rows:
        _apply_header_style(ws, ncols)

    # Explicit column widths win over autofit.
    widths = sheet_spec.get("column_widths") or {}
    if widths and isinstance(widths, dict):
        for col, w in widths.items():
            try:
                ws.column_dimensions[str(col)].width = float(w)
            except (ValueError, TypeError):
                pass
    elif rows:
        _autofit_columns(ws, [r for r in rows if isinstance(r, (list, tuple))])

    freeze = sheet_spec.get("freeze_panes")
    if freeze:
        ws.freeze_panes = str(freeze)

    cf = sheet_spec.get("conditional_format")
    if isinstance(cf, dict):
        _apply_conditional_format(ws, cf)

    # 数字格式(货币/百分比/千分位/日期) → 逐格样式(底色/字色/边框/合并)。
    # 顺序: number_format 先,merges 后(让合并锚点带样式),cell_styles 最后
    # 覆盖(用户显式样式优先级最高)。
    nrows = len(rows)
    _apply_number_formats(ws, sheet_spec.get("number_formats"), nrows)
    _apply_merges(ws, sheet_spec.get("merge_cells"))
    _apply_cell_styles(ws, sheet_spec.get("cell_styles"))

    # 单图表(BC)+ 多图表(charts 列表)。
    chart = sheet_spec.get("chart")
    if isinstance(chart, dict):
        try:
            _add_chart(ws, chart)
        except Exception as exc:  # noqa: BLE001 — chart is non-critical
            log.warning("excel chart skipped: %s", exc)
    charts = sheet_spec.get("charts")
    if isinstance(charts, list):
        for cs in charts:
            if isinstance(cs, dict):
                try:
                    _add_chart(ws, cs)
                except Exception as exc:  # noqa: BLE001
                    log.warning("excel chart skipped: %s", exc)

    _add_images(ws, sheet_spec.get("images"))


def excel_create(
    spec: Any,
    *,
    output_path: Optional[str] = None,
) -> dict[str, Any]:
    """Render ``spec`` into a ``.xlsx``. See module docstring for shape.

    Returns ``{"ok": True, "path", "sheet_count"}`` or
    ``{"ok": False, "error": ...}``.
    """
    if not _HAS_OPENPYXL:
        return {
            "ok": False,
            "error": "openpyxl not installed; install with `pip install openpyxl`",
            "retriable": False,
        }

    spec_dict = _coerce_spec(spec)
    if spec_dict is None:
        return {"ok": False, "error": "spec must be a dict or JSON object string", "retriable": False}

    sheets = spec_dict.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        return {"ok": False, "error": "spec.sheets must be a non-empty list", "retriable": False}

    try:
        out_path = office_paths.resolve_for_write(
            output_path, default_prefix="deskpet-excel", default_suffix=".xlsx",
            default_kind="Excel",
        )
    except office_paths.PathError as exc:
        return {"ok": False, "error": str(exc), "retriable": False}

    try:
        wb = Workbook()
        for i, sheet_spec in enumerate(sheets):
            if not isinstance(sheet_spec, dict):
                continue
            _build_sheet(wb, sheet_spec, first=(i == 0))
        wb.save(str(out_path))
    except Exception as exc:  # noqa: BLE001
        log.warning("excel_create failed: %s", exc, exc_info=True)
        return {"ok": False, "error": f"render failed: {exc}", "retriable": True}

    # WI-T1.2 D1：显式 emit artifacts[]（一等公民路径，保 BC）
    return {
        "ok": True,
        "path": str(out_path),
        "sheet_count": len(sheets),
        "artifacts": [{
            "kind": "file",
            "path": str(out_path),
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "title": Path(str(out_path)).name,
        }],
    }


# ---------------------------------------------------------------------------
# Tool registry wiring
# ---------------------------------------------------------------------------
_SCHEMA = {
    "name": "excel_create",
    "description": (
        "Generate a professional .xlsx spreadsheet locally from a structured "
        "spec. Supports multiple sheets, live formulas (any cell value "
        "starting with '='), charts (bar/line/pie, single or many per sheet), "
        "conditional formatting, header styling, frozen panes, column widths, "
        "number formats (currency/percent/thousands/date), merged cells, "
        "per-cell styling (fill/font color/border/align) and embedded images. "
        "Returns the file path (default 桌宠/OutPut/Excel) — tell the user the "
        "full path. Decide the data layout first, then call this."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "spec": {
                "description": (
                    "Dict (or JSON string of a dict) with a 'sheets' list. Each "
                    "sheet supports: name, rows (list of row lists), header_row, "
                    "column_widths, freeze_panes, conditional_format, chart, "
                    "charts (list), number_formats (list of {range|col, format} "
                    "e.g. {col:'B',format:'#,##0.00'} or {col:'C',format:'0.0%'}),"
                    " merge_cells (list of 'A1:C1'), cell_styles (list of "
                    "{range, bold, italic, font_color:'#RRGGBB', fill:'#RRGGBB', "
                    "align, font_size, border, wrap}), images (list of {path, "
                    "anchor, width, height}). A cell value like '=SUM(B2:B5)' "
                    "becomes a live formula."
                ),
            },
            "output_path": {
                "type": "string",
                "description": (
                    "Absolute .xlsx path. Omit to auto-save under "
                    "桌宠/OutPut/Excel. A non-temp directory must have been "
                    "picked by the user."
                ),
            },
        },
        "required": ["spec"],
    },
}


def _handle(args: dict[str, Any], task_id: str = "") -> str:
    result = excel_create(
        args.get("spec"),
        output_path=(str(args["output_path"]) if args.get("output_path") else None),
    )
    return json.dumps(result, ensure_ascii=False)


def _register() -> None:
    try:
        from .registry import registry

        registry.register(
            "excel_create",
            "office",
            _SCHEMA,
            _handle,
            permission_category="write_file",
            timeout_seconds=30.0,
            concurrency_safe=False,  # G3: writes .xlsx to disk
        )
    except Exception:  # noqa: BLE001
        pass


_register()

__all__ = ["excel_create"]
