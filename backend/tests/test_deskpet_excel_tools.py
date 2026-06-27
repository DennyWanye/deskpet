# SPDX-FileCopyrightText: 2026 DennyWanye
# SPDX-License-Identifier: BUSL-1.1

"""TDD T1 — excel_tools.py (excel_create)."""
from __future__ import annotations

from pathlib import Path

import pytest

from deskpet.tools import office_paths as op
from deskpet.tools import excel_tools as xl

pytestmark = pytest.mark.skipif(not xl._HAS_OPENPYXL, reason="openpyxl not installed")


@pytest.fixture(autouse=True)
def _auth(tmp_path: Path):
    op.clear_authorizations()
    op.authorize_path(tmp_path)
    yield
    op.clear_authorizations()


def _load(path: str):
    from openpyxl import load_workbook

    return load_workbook(path)


def test_t1_1_single_sheet_data(tmp_path: Path):
    out = tmp_path / "a.xlsx"
    spec = {"sheets": [{"name": "S1", "rows": [["x", "y"], [1, 2]]}]}
    r = xl.excel_create(spec, output_path=str(out))
    assert r["ok"], r
    wb = _load(r["path"])
    assert wb["S1"]["A2"].value == 1


def test_t1_2_multi_sheet(tmp_path: Path):
    spec = {"sheets": [
        {"name": "Alpha", "rows": [["a"]]},
        {"name": "Beta", "rows": [["b"]]},
    ]}
    r = xl.excel_create(spec, output_path=str(tmp_path / "m.xlsx"))
    assert r["ok"]
    wb = _load(r["path"])
    assert wb.sheetnames == ["Alpha", "Beta"]


def test_t1_3_formula_cell(tmp_path: Path):
    spec = {"sheets": [{"name": "S", "rows": [[1], [2], ["=SUM(A1:A2)"]]}]}
    r = xl.excel_create(spec, output_path=str(tmp_path / "f.xlsx"))
    assert r["ok"]
    wb = _load(r["path"])
    assert str(wb["S"]["A3"].value).startswith("=SUM")


def test_t1_4_chart_present(tmp_path: Path):
    spec = {"sheets": [{
        "name": "S",
        "rows": [["cat", "val"], ["a", 3], ["b", 5]],
        "chart": {"type": "bar", "title": "T", "data": "B1:B3", "categories": "A2:A3", "anchor": "D2"},
    }]}
    r = xl.excel_create(spec, output_path=str(tmp_path / "c.xlsx"))
    assert r["ok"]
    wb = _load(r["path"])
    assert len(wb["S"]._charts) >= 1


def test_t1_5_conditional_format(tmp_path: Path):
    spec = {"sheets": [{
        "name": "S",
        "rows": [["v"], [1], [9]],
        "conditional_format": {"range": "A2:A3", "rule": "color_scale"},
    }]}
    r = xl.excel_create(spec, output_path=str(tmp_path / "cf.xlsx"))
    assert r["ok"]
    wb = _load(r["path"])
    assert len(list(wb["S"].conditional_formatting)) >= 1


def test_t1_6_header_style_and_width(tmp_path: Path):
    spec = {"sheets": [{
        "name": "S",
        "rows": [["Head1", "Head2"], [1, 2]],
        "header_row": True,
        "column_widths": {"A": 25},
    }]}
    r = xl.excel_create(spec, output_path=str(tmp_path / "h.xlsx"))
    assert r["ok"]
    wb = _load(r["path"])
    ws = wb["S"]
    assert ws["A1"].font.bold
    assert ws.column_dimensions["A"].width == 25


def test_t1_7_spec_as_json_string(tmp_path: Path):
    spec = '{"sheets": [{"name": "S", "rows": [["a", 1]]}]}'
    r = xl.excel_create(spec, output_path=str(tmp_path / "j.xlsx"))
    assert r["ok"]


def test_t1_8_bad_spec_no_crash():
    r = xl.excel_create({"not_sheets": []})
    assert not r["ok"] and r["retriable"] is False


def test_t1_9_no_output_path_goes_to_temp():
    r = xl.excel_create({"sheets": [{"name": "S", "rows": [["a"]]}]})
    assert r["ok"]
    assert Path(r["path"]).exists()
    Path(r["path"]).unlink(missing_ok=True)


def test_t1_10_system_path_refused():
    r = xl.excel_create(
        {"sheets": [{"name": "S", "rows": [["a"]]}]},
        output_path="C:\\Windows\\evil.xlsx",
    )
    assert not r["ok"] and r["retriable"] is False


def test_t1_11_empty_sheets_list_rejected():
    r = xl.excel_create({"sheets": []})
    assert not r["ok"]


# --- 复杂 Excel 升级 (number_formats / merge / cell_styles / charts / images) ---

def test_t1_12_number_format_by_col(tmp_path: Path):
    spec = {"sheets": [{
        "name": "钱",
        "rows": [["项目", "金额"], ["A", 1200], ["B", 300]],
        "header_row": True,
        "number_formats": [{"col": "B", "format": "#,##0.00"}],
    }]}
    r = xl.excel_create(spec, output_path=str(tmp_path / "nf.xlsx"))
    assert r["ok"], r
    wb = _load(r["path"])
    ws = wb["钱"]
    assert ws["B2"].number_format == "#,##0.00"
    assert ws["B3"].number_format == "#,##0.00"
    # 表头行(row1)不应被列格式影响(from_row 默认 2)
    assert ws["B1"].number_format != "#,##0.00"


def test_t1_13_number_format_by_range(tmp_path: Path):
    spec = {"sheets": [{
        "name": "P",
        "rows": [["率"], [0.25], [0.5]],
        "number_formats": [{"range": "A2:A3", "format": "0.0%"}],
    }]}
    r = xl.excel_create(spec, output_path=str(tmp_path / "pct.xlsx"))
    assert r["ok"], r
    ws = _load(r["path"])["P"]
    assert ws["A2"].number_format == "0.0%"


def test_t1_14_merge_cells(tmp_path: Path):
    spec = {"sheets": [{
        "name": "M",
        "rows": [["季度汇总", None, None], ["Q1", "Q2", "Q3"]],
        "merge_cells": ["A1:C1"],
    }]}
    r = xl.excel_create(spec, output_path=str(tmp_path / "mg.xlsx"))
    assert r["ok"], r
    ws = _load(r["path"])["M"]
    assert "A1:C1" in [str(rng) for rng in ws.merged_cells.ranges]


def test_t1_15_cell_styles_fill_and_color(tmp_path: Path):
    spec = {"sheets": [{
        "name": "St",
        "rows": [["标题"], ["数据"]],
        "cell_styles": [{
            "range": "A1",
            "bold": True,
            "fill": "#1F4E78",
            "font_color": "#FFFFFF",
            "align": "center",
            "border": True,
        }],
    }]}
    r = xl.excel_create(spec, output_path=str(tmp_path / "st.xlsx"))
    assert r["ok"], r
    ws = _load(r["path"])["St"]
    cell = ws["A1"]
    assert cell.font.bold is True
    assert cell.fill.fgColor.rgb == "FF1F4E78"
    assert cell.font.color.rgb == "FFFFFFFF"
    assert cell.border.left.style == "thin"


def test_t1_16_multiple_charts(tmp_path: Path):
    spec = {"sheets": [{
        "name": "Ch",
        "rows": [["m", "v"], ["a", 1], ["b", 2]],
        "charts": [
            {"type": "bar", "data": "B1:B3", "categories": "A2:A3", "anchor": "D2"},
            {"type": "line", "data": "B1:B3", "categories": "A2:A3", "anchor": "D20"},
        ],
    }]}
    r = xl.excel_create(spec, output_path=str(tmp_path / "ch.xlsx"))
    assert r["ok"], r
    ws = _load(r["path"])["Ch"]
    assert len(ws._charts) == 2


def test_t1_17_image_embedded(tmp_path: Path):
    if not xl._HAS_XLIMAGE:
        pytest.skip("openpyxl image support / Pillow missing")
    # 造一张最小 PNG
    from PIL import Image as _PILImage

    img_path = tmp_path / "dot.png"
    _PILImage.new("RGB", (10, 10), (255, 0, 0)).save(img_path)
    spec = {"sheets": [{
        "name": "Img",
        "rows": [["see chart"]],
        "images": [{"path": str(img_path), "anchor": "C2"}],
    }]}
    r = xl.excel_create(spec, output_path=str(tmp_path / "img.xlsx"))
    assert r["ok"], r
    ws = _load(r["path"])["Img"]
    assert len(ws._images) == 1


def test_t1_18_default_path_under_output_excel(tmp_path: Path, monkeypatch):
    # default_kind='Excel' → 落 <user_data>/OutPut/Excel (这里把 user_data 指到 tmp)
    monkeypatch.setenv("DESKPET_USER_DATA_DIR", str(tmp_path))
    r = xl.excel_create({"sheets": [{"name": "S", "rows": [["a"]]}]})
    assert r["ok"], r
    p = Path(r["path"])
    assert p.exists()
    assert p.parent.name == "Excel"
    assert p.parent.parent.name == "OutPut"
    p.unlink(missing_ok=True)
